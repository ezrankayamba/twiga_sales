import openpyxl
from . import models
from zipfile import ZipFile
from django.core.files import File
from io import TextIOWrapper
import re
import io
from django.core.files.base import ContentFile
import threading
from . import ocr, ocr2
import xlsxwriter
import json
import traceback
from . import serializers
from websocket import alarms


def doc_key(name):
    return f'{name.lower()}_doc'


def docs_schema():
    # return serializers.SchemaSerializer(models.Schema.objects.all(), many=True).data
    return [
        {'name': models.Document.DOC_C2, 'letter': models.Document.LETTER_C2, 'key': doc_key(
            models.Document.DOC_C2), 'regex': '[ ]{0,1}(T\w{14,})[\({ ]', 'params': {'x': 700, 'y': 600, 'h': 400, 'w': 800, 'threshold': 225}, 'mandatory': True, 'corrections': [{'pos': 1, 'val': '2', 'rep': 'Z'}]},
        {'name': models.Document.DOC_ASSESSMENT, 'prefix': 'C ', 'letter': models.Document.LETTER_ASSESSMENT, 'key': doc_key(models.Document.DOC_ASSESSMENT), 'regex': '[CcG]{1,}[ ]{0,}(\d{2,})', 'params': {
            'x': 700, 'y': 20, 'h': 500, 'w': 800, 'threshold': 230, 'zoom': 1.1}, 'mandatory': True},
        {'name': models.Document.DOC_EXIT, 'letter': models.Document.LETTER_EXIT, 'key': doc_key(
            models.Document.DOC_EXIT), 'regex': 'laration[ +\d:]{1,}(\d{4} [\w/]+)', 'params': {'x': 20, 'y': 600, 'h': 600, 'w': 600, 'threshold': 236}, 'mandatory': False}
    ]


def import_sales(batch):
    excel_file = batch.file_in
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    i = 0
    rows = []
    for row in ws.values:
        if i:
            res = {'TRANS DATE': '', 'CUSTOMER': '', 'DELIVERY NOTE': '', 'VEH#': '',
                   'TAX INVOICE': '', 'SO#': '', 'PRODUCT': '', 'QTY9TONS': '', 'VALUE': '', 'DESTINATION': '', 'STATUS': '', 'DETAILS': ''}
            dict = {}
            try:
                dict['transaction_date'] = row[0]
                res['TRANS DATE'] = row[0]
                dict['customer_name'] = row[1]
                res['CUSTOMER'] = row[1]
                dict['delivery_note'] = row[2]
                res['DELIVERY NOTE'] = row[2]
                dict['vehicle_number'] = row[3]
                res['VEH#'] = row[3]
                dict['tax_invoice'] = row[4]
                res['TAX INVOICE'] = row[4]
                dict['sales_order'] = row[5]
                res['SO#'] = row[5]
                dict['product_name'] = row[6]
                res['PRODUCT'] = row[6]
                dict['quantity'] = row[7]
                res['QTY9TONS'] = row[7]
                dict['total_value'] = row[8]
                res['VALUE'] = row[8]
                dict['destination'] = row[9]
                res['DESTINATION'] = row[9]
                dict['vehicle_number_trailer'] = row[10] if len(row) > 10 else None
                res['VEH# TRAILER'] = row[10] if len(row) > 10 else None
                models.Sale.objects.create(**dict)
                res['STATUS'] = 'Success'
                res['DETAILS'] = json.dumps({'errors': []})
                rows.append(res)
            except Exception as e:
                print(e)
                err_msg = str(e)
                if 'UNIQUE' in err_msg:
                    err_msg = 'Duplicate record'
                res['STATUS'] = 'Fail'
                res['DETAILS'] = json.dumps({'errors': [{'message': err_msg}]})
                rows.append(res)
        else:
            if len(row) < 2:
                print('Not enough columns')
                break
        i += 1

    write_out(batch, rows, headers=['TRANS DATE', 'CUSTOMER', 'DELIVERY NOTE', 'VEH#',
                                    'TAX INVOICE', 'SO#', 'PRODUCT', 'QTY9TONS', 'VALUE', 'DESTINATION', 'STATUS', 'DETAILS', 'VEH# TRAILER'])
    print('Completed processing the upload')


def map_error(error):
    res = {}
    key = error['name']
    res[key] = error['message']
    return res


def read_entries(zip, row, docs_list, agent):
    so, qty, val = (row[0], row[1], row[2])
    sale = models.Sale.objects.filter(sales_order=so.strip()).first() if so else None
    if sale:
        truck = 'trailer' if sale.quantity >= models.TRUCK_THRESHOLD else 'head'
        doc_entries = filter(lambda x: so in x.filename, docs_list)
        errors, docs = ([], [])
        for doc_entry in doc_entries:
            with zip.open(doc_entry, 'r') as file:
                filename = doc_entry.filename.split("/")[1]
                doc_type = filename[0]
                d = None
                try:
                    d = next(x for x in docs_schema() if doc_type == x['letter'])
                except Exception as e:
                    print(e)
                    res = {'sale': sale, 'result': -1, 'errors': [{'sale': 'Sales order number is not valid'}]}
                    errors.append({
                        'key': 'Unknown',
                        'name': "Unknown",
                        'message': "Not valid document or invalid prefix",
                        'mandatory': True
                    })
                    continue

                args, name, regex, pdf_data = (d['params'], d['name'], d['regex'], io.BytesIO(file.read()))
                print(name, f'"{regex}"')
                # ref_number = ocr.new_extract_from_file(regex, pdf_data, **args)
                ref_number = ocr2.extract_ref_number(pdf_data, regex,  **args)
                print("Resulting Ref No: ", ref_number)
                error = None
                if ref_number:
                    if 'corrections' in d:
                        ref_number = ocr2.apply_corrections(ref_number, d['corrections'])
                    prefix = d.get('prefix', '')
                    ref_number = f'{prefix}{ref_number}'
                    print(ref_number)
                    duplicate = models.Document.objects.filter(ref_number=ref_number, truck=truck).first()
                    if duplicate:
                        error = f'Duplicate {name} document'
                    else:
                        docs.append({
                            'ref_number': ref_number,
                            'file': File(pdf_data, name=filename),
                            'sale': sale,
                            'doc_type': name,
                            'truck': truck,
                            'user': agent.user
                        })
                else:
                    error = f'Invalid {name} document'
                if error:
                    errors.append({
                        'key': d['key'],
                        'name': name,
                        'message': error,
                        'mandatory': d['mandatory']
                    })

        if len(list(filter(lambda x: x['mandatory'], errors))):
            print('Errors: ', errors)
        else:
            print('Docs: ', docs)
            sale.agent = agent
            sale.quantity2 = qty
            sale.total_value2 = val
            sale.save()
            for doc in docs:
                models.Document.objects.create(**doc)
        return {'sale': sale, 'result': 0, 'errors': list(map(map_error, errors))}
    else:
        return {'sale': sale, 'result': -1, 'errors': [{'sale': 'Sales order number is not valid'}]}


def cell(i, j):
    char = "A"
    char = chr(ord(char[0]) + j - 1)
    return f'{char}{i}'


def write_out(batch, rows, headers):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    main = workbook.add_worksheet("Result")
    print(rows[0] if len(rows) else rows, headers)

    for j, col in enumerate(headers, start=1):
        main.write(f'{cell(1, j)}', col)

    for i, row in enumerate(rows, start=2):
        for j, col in enumerate(row, start=1):
            main.write(f'{cell(i, j)}', row[col])

    workbook.close()
    xlsx_data = output.getvalue()
    batch.file_out = File(output, name=f'Result_{batch.id}.xlsx')
    # batch.file_out.save()
    batch.status = 1
    batch.save()


def import_docs(batch):
    agent = batch.user.agent
    rows = []
    with ZipFile(batch.file_in, 'r') as zip:
        docs_list = []
        for entry in zip.infolist():
            if '/' in entry.filename:
                docs_list.append(entry)
            else:
                excel = entry
        sales = []
        if excel:
            with zip.open(excel, 'r') as file:
                ws = openpyxl.load_workbook(ContentFile(file.read())).active
                i = 0

                for row in ws.values:
                    print("Data ........", row)
                    if i and len(row) >= 3:
                        print(row)
                        rec = {}
                        try:
                            rec['SO#'] = row[0]
                            rec['Quantity'] = row[1]
                            rec['Volume'] = row[2]
                            res = read_entries(zip, row, docs_list, agent)
                            rec['Status'] = 'Completed' if res['result'] == 0 else 'Failed'
                            rec['Detail'] = json.dumps({'errors': res['errors']})
                        except Exception as e:
                            print("Error", e)
                            traceback.print_exc()
                            rec['Status'] = 'Failed'
                            rec['Detail'] = json.dumps({'errors': [{'message': f'Exception: {e}'}]})
                        rows.append(rec)
                    i += 1

    headers = ['SO#', 'Quantity', 'Volume', 'Status', 'Detail']
    write_out(batch, rows, headers=headers)
    data = serializers.BatchSerializer(batch).data
    print(data)
    alarms.trigger(batch.user, {'file_in': batch.file_in.url, 'file_out': batch.file_out.url})
    print('Completed processing the upload')


def sales_import_async(batch):
    t = threading.Thread(target=import_sales, args=(batch,), kwargs={})
    t.setDaemon(True)
    t.start()


def docs_import_async(batch):
    t = threading.Thread(target=import_docs, args=(batch,), kwargs={})
    t.setDaemon(True)
    t.start()
